import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def obtener_dia_semana():
    """Solicita al usuario el número del día de la semana"""
    dias = {1: "Lunes", 2: "Martes", 3: "Miércoles", 4: "Jueves", 5: "Viernes", 6: "Sábado", 7: "Domingo"}
    
    while True:
        try:
            print("\n=== Seleccione el día de la semana ===")
            for num, dia in dias.items():
                print(f"{num}. {dia}")
            dia_num = int(input("\nIngrese el número del día (1-7): "))
            if dia_num in dias:
                return dias[dia_num]
            print("Por favor, ingrese un número entre 1 y 7")
        except ValueError:
            print("Error: Por favor ingrese un número válido")

def buscar_archivos(ruta_carpeta):
    """Busca los archivos de Plantilla y Asistencia"""
    try:
        archivos = os.listdir(ruta_carpeta)
        archivo_plantilla, archivo_asistencia = None, None
        
        for archivo in archivos:
            if archivo.startswith('~$'):
                continue
            if archivo.endswith(('.xlsx', '.xls')):
                if 'Plantilla' in archivo:
                    archivo_plantilla = os.path.join(ruta_carpeta, archivo)
                elif 'Asistencia' in archivo:
                    archivo_asistencia = os.path.join(ruta_carpeta, archivo)
        
        return archivo_plantilla, archivo_asistencia
    except FileNotFoundError:
        print(f"Error: No se encontró la carpeta {ruta_carpeta}")
        return None, None

def leer_plantilla(archivo_plantilla):
    """Lee el archivo de plantilla y encuentra la hoja correcta"""
    xls = pd.ExcelFile(archivo_plantilla)
    hoja = None
    
    print(f"\nHojas en Plantilla: {xls.sheet_names}")
    for h in xls.sheet_names:
        if 'Rutero' in h or 'Promotoria' in h:
            hoja = h
            break
    
    if not hoja:
        print("\nNo se encontró hoja automáticamente. Seleccione:")
        for i, h in enumerate(xls.sheet_names, 1):
            print(f"{i}. {h}")
        sel = int(input("Número de hoja: ")) - 1
        hoja = xls.sheet_names[sel]
    
    print(f"Usando hoja: {hoja}")
    
    # Buscar fila de encabezado
    df_preview = pd.read_excel(archivo_plantilla, sheet_name=hoja, header=None, nrows=10)
    print("\nPrimeras filas:")
    print(df_preview)
    
    fila_enc = None
    for i in range(10):
        try:
            fila_str = ' '.join(df_preview.iloc[i].astype(str).tolist())
            if 'Lunes' in fila_str or 'Martes' in fila_str or 'Miércoles' in fila_str:
                fila_enc = i
                print(f"\nEncabezados en fila {i}")
                break
        except:
            continue
    
    if fila_enc is None:
        fila_enc = int(input("Ingrese fila de encabezados: "))
    
    return pd.read_excel(archivo_plantilla, sheet_name=hoja, header=fila_enc)

def leer_asistencia(archivo_asistencia):
    """Lee el archivo de asistencia y encuentra la hoja correcta"""
    xls = pd.ExcelFile(archivo_asistencia)
    hoja = None
    
    print(f"\nHojas en Asistencia: {xls.sheet_names}")
    for h in xls.sheet_names:
        if 'Encuesta' in h or 'Asistencia' in h or 'Reporte' in h:
            hoja = h
            break
    
    if not hoja:
        print("\nNo se encontró hoja automáticamente. Seleccione:")
        for i, h in enumerate(xls.sheet_names, 1):
            print(f"{i}. {h}")
        sel = int(input("Número de hoja: ")) - 1
        hoja = xls.sheet_names[sel]
    
    print(f"Usando hoja: {hoja}")
    return pd.read_excel(archivo_asistencia, sheet_name=hoja)

def generar_reporte(df_plantilla, df_asistencia, dia, ruta_salida):
    """Genera el reporte en formato ejecutivo"""
    
    # Convertir columna del día a numérico
    df_plantilla[dia] = pd.to_numeric(df_plantilla[dia], errors='coerce')
    
    # Limpiar espacios
    df_plantilla['Nombre del Supervisor'] = df_plantilla['Nombre del Supervisor'].astype(str).str.strip()
    df_plantilla['Nombre del Promotor'] = df_plantilla['Nombre del Promotor'].astype(str).str.strip()
    df_plantilla['Sucursal'] = df_plantilla['Sucursal'].astype(str).str.strip()
    df_plantilla['Ruta'] = df_plantilla['Ruta'].astype(str).str.strip()
    df_asistencia['Encuestador/Tienda'] = df_asistencia['Encuestador/Tienda'].astype(str).str.strip()
    
    # Sucursales visitadas
    sucursales_visitadas = set(df_asistencia['Encuestador/Tienda'].dropna().unique())
    print(f"\nSucursales visitadas encontradas: {len(sucursales_visitadas)}")
    
    # Filtrar tiendas del día (número >= 1)
    df_dia = df_plantilla[(df_plantilla[dia].notna()) & (df_plantilla[dia] >= 1)].copy()
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {dia}"
    
    # Estilos
    font_bold = Font(bold=True)
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    fila_actual = 1
    
    # === RESUMEN POR SUPERVISOR ===
    ws.cell(row=fila_actual, column=1, value="Supervisor").font = font_bold
    ws.cell(row=fila_actual, column=2, value="Objetivo").font = font_bold
    ws.cell(row=fila_actual, column=3, value="Ajuste").font = font_bold
    ws.cell(row=fila_actual, column=4, value="Efectividad con ajuste").font = font_bold
    fila_actual += 1
    
    supervisores = df_dia['Nombre del Supervisor'].unique()
    resumen_supervisores = {}
    
    for sup in supervisores:
        if sup == 'nan' or pd.isna(sup):
            continue
        df_sup = df_dia[df_dia['Nombre del Supervisor'] == sup]
        objetivo = len(df_sup)
        
        # Contar visitas
        visitas = sum(1 for _, row in df_sup.iterrows() if row['Sucursal'] in sucursales_visitadas)
        efectividad = round((visitas / objetivo * 100), 0) if objetivo > 0 else 0
        
        resumen_supervisores[sup] = {'objetivo': objetivo, 'visitas': visitas, 'efectividad': efectividad}
        
        ws.cell(row=fila_actual, column=1, value=sup)
        ws.cell(row=fila_actual, column=2, value=objetivo)
        ws.cell(row=fila_actual, column=3, value=visitas)
        ws.cell(row=fila_actual, column=4, value=f"{efectividad}%")
        fila_actual += 1
    
    fila_actual += 2
    
    # === DETALLE POR SUPERVISOR ===
    for sup in supervisores:
        if sup == 'nan' or pd.isna(sup):
            continue
        
        df_sup = df_dia[df_dia['Nombre del Supervisor'] == sup]
        
        # Calcular totales del supervisor
        objetivo_total = len(df_sup)
        efectividad_total = sum(1 for _, row in df_sup.iterrows() if row['Sucursal'] in sucursales_visitadas)
        
        # Encabezado de totales
        ws.cell(row=fila_actual, column=5, value="Objetivo de Visitas").font = font_bold
        ws.cell(row=fila_actual, column=6, value="Efectividad Total").font = font_bold
        fila_actual += 1
        ws.cell(row=fila_actual, column=5, value=objetivo_total)
        ws.cell(row=fila_actual, column=6, value=efectividad_total)
        fila_actual += 1
        
        # Encabezados de detalle
        headers = ["Nombre del Supervisor", "Ruta", "Nombre del Promotor", "Sucursal", f"Objetivo {dia}", "Visitas Ejecutadas"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=fila_actual, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = border
        fila_actual += 1
        
        # Datos por promotor
        for _, row in df_sup.iterrows():
            sucursal = row['Sucursal']
            visitada = 1 if sucursal in sucursales_visitadas else 0
            
            # Emoji de palomita o X según si fue visitada
            estado = "✓" if visitada else "✗"
            
            ws.cell(row=fila_actual, column=1, value=sup).border = border
            ws.cell(row=fila_actual, column=2, value=row['Ruta']).border = border
            ws.cell(row=fila_actual, column=3, value=row['Nombre del Promotor']).border = border
            ws.cell(row=fila_actual, column=4, value=sucursal).border = border
            ws.cell(row=fila_actual, column=5, value=1).border = border  # Objetivo siempre es 1
            
            cell_visita = ws.cell(row=fila_actual, column=6, value=estado)
            cell_visita.border = border
            cell_visita.alignment = Alignment(horizontal='center')
            
            # Resaltar en verde las visitadas y rojo las no visitadas
            if visitada:
                cell_visita.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell_visita.font = Font(color="006100", bold=True)
            else:
                cell_visita.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell_visita.font = Font(color="9C0006", bold=True)
            
            fila_actual += 1
        
        fila_actual += 2  # Espacio entre supervisores
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    # Guardar
    nombre_archivo = f"Reporte de Cierre de Efectividad {dia}.xlsx"
    ruta_completa = os.path.join(ruta_salida, nombre_archivo)
    wb.save(ruta_completa)
    
    return ruta_completa, resumen_supervisores, nombre_archivo

def crear_carpeta_y_mover_archivos(ruta_base, dia, archivo_plantilla, archivo_asistencia, archivo_reporte):
    """Crea una carpeta con el día y fecha, y mueve los archivos utilizados y generados"""
    from datetime import datetime
    import shutil
    
    # Obtener fecha actual
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    
    # Crear nombre de carpeta
    nombre_carpeta = f"Cierre Efectividad {dia} {fecha_actual}"
    ruta_carpeta = os.path.join(ruta_base, nombre_carpeta)
    
    # Crear carpeta si no existe
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)
        print(f"\n✓ Carpeta creada: {nombre_carpeta}")
    
    # Mover archivos
    archivos_a_mover = [
        (archivo_plantilla, os.path.basename(archivo_plantilla)),
        (archivo_asistencia, os.path.basename(archivo_asistencia)),
        (archivo_reporte, os.path.basename(archivo_reporte))
    ]
    
    for origen, nombre in archivos_a_mover:
        destino = os.path.join(ruta_carpeta, nombre)
        try:
            shutil.move(origen, destino)
            print(f"  ✓ Movido: {nombre}")
        except Exception as e:
            print(f"  ✗ Error al mover {nombre}: {e}")
    
    return ruta_carpeta

def main():
    print("=" * 60)
    print("SISTEMA DE VERIFICACIÓN DE ASISTENCIA Y PLANTILLA")
    print("=" * 60)
    
    dia = obtener_dia_semana()
    ruta = r"C:\Users\lapmxdf558\Documents\Archivos Alejandro\Genomma Mayoreo\Asistencia\Efectividad Automatico"
    
    print(f"\nBuscando archivos en: {ruta}")
    
    archivo_plantilla, archivo_asistencia = buscar_archivos(ruta)
    
    if not archivo_plantilla:
        print("\nError: No se encontró archivo de Plantilla")
        return
    if not archivo_asistencia:
        print("\nError: No se encontró archivo de Asistencia")
        return
    
    print(f"\nPlantilla: {os.path.basename(archivo_plantilla)}")
    print(f"Asistencia: {os.path.basename(archivo_asistencia)}")
    
    try:
        df_plantilla = leer_plantilla(archivo_plantilla)
        df_asistencia = leer_asistencia(archivo_asistencia)
        
        # Verificar columna del día
        if dia not in df_plantilla.columns:
            print(f"\nColumnas disponibles: {df_plantilla.columns.tolist()}")
            raise ValueError(f"No se encontró la columna '{dia}'")
        
        # Generar reporte
        ruta_archivo, resumen, nombre_reporte = generar_reporte(df_plantilla, df_asistencia, dia, ruta)
        
        print("\n" + "=" * 60)
        print(f"RESUMEN DE EFECTIVIDAD - {dia.upper()}")
        print("=" * 60)
        
        total_obj, total_vis = 0, 0
        for sup, datos in resumen.items():
            print(f"\n{sup}")
            print(f"  Objetivo: {datos['objetivo']} | Visitas: {datos['visitas']} | Efectividad: {datos['efectividad']}%")
            total_obj += datos['objetivo']
            total_vis += datos['visitas']
        
        print("\n" + "-" * 40)
        print(f"TOTAL: Objetivo {total_obj} | Visitas {total_vis} | Efectividad {round(total_vis/total_obj*100, 1) if total_obj > 0 else 0}%")
        
        print(f"\n✓ Reporte guardado: {ruta_archivo}")
        
        # Crear carpeta y mover archivos
        print("\n" + "=" * 60)
        print("ORGANIZANDO ARCHIVOS")
        print("=" * 60)
        ruta_carpeta_final = crear_carpeta_y_mover_archivos(ruta, dia, archivo_plantilla, archivo_asistencia, ruta_archivo)
        print(f"\n✓ Archivos organizados en: {ruta_carpeta_final}")
        
    except Exception as e:
        print(f"\nError: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()