import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta

def generar_template_tiendas(archivo_entrada, numero_promotor, fecha_inicio, fecha_fin, archivo_salida='Template.xlsx'):
    """
    Genera un template de Excel con las columnas: Fecha, Promotor, Tienda
    Repite todas las tiendas para cada día en el rango de fechas especificado.
    
    Parámetros:
    - archivo_entrada: Ruta al archivo Excel con la lista de tiendas
    - numero_promotor: Número del promotor a asignar
    - fecha_inicio: Fecha de inicio en formato 'DD/MM/YYYY' o 'YYYY-MM-DD'
    - fecha_fin: Fecha fin en formato 'DD/MM/YYYY' o 'YYYY-MM-DD'
    - archivo_salida: Nombre del archivo de salida (default: 'Template.xlsx')
    """
    
    try:
        # Leer el archivo de entrada
        df = pd.read_excel(archivo_entrada)
        
        # Mostrar información del archivo
        print(f"📊 Columnas encontradas: {len(df.columns)}")
        print(f"📋 Nombres de columnas: {list(df.columns)}")
        print(f"📝 Total de filas: {len(df)}\n")
        
        # Convertir fechas
        if '/' in fecha_inicio:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%d/%m/%Y')
            fecha_fin_dt = datetime.strptime(fecha_fin, '%d/%m/%Y')
        else:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
        
        # Generar lista de fechas
        fechas = []
        fecha_actual = fecha_inicio_dt
        while fecha_actual <= fecha_fin_dt:
            fechas.append(fecha_actual)
            fecha_actual += timedelta(days=1)
        
        print(f"📅 Generando template del {fecha_inicio_dt.strftime('%d/%m/%Y')} al {fecha_fin_dt.strftime('%d/%m/%Y')}")
        print(f"📅 Total de días: {len(fechas)}\n")
        
        # Determinar qué columna tiene las tiendas
        # Buscar la columna que parece tener nombres de tiendas
        columna_tienda = None
        
        for col in df.columns:
            # Verificar si la columna tiene texto largo (probablemente nombres de tiendas)
            muestra = df[col].dropna().head(5)
            if muestra.dtype == 'object' and len(muestra) > 0:
                # Verificar si contiene nombres de tiendas (texto largo)
                promedio_longitud = muestra.astype(str).str.len().mean()
                if promedio_longitud > 10:  # Nombres de tiendas suelen ser largos
                    columna_tienda = col
                    break
        
        if columna_tienda is None:
            # Si no encuentra automáticamente, usar la última columna
            columna_tienda = df.columns[-1]
        
        print(f"🏪 Usando columna '{columna_tienda}' para las tiendas")
        
        # Obtener lista única de tiendas
        tiendas = df[columna_tienda].dropna().unique()
        print(f"🏪 Total de tiendas únicas: {len(tiendas)}\n")
        
        # Crear el DataFrame para el template
        template_data = []
        
        for fecha in fechas:
            for tienda in tiendas:
                template_data.append({
                    'Fecha': fecha,
                    'Promotor': numero_promotor,
                    'Tienda': str(tienda).strip()
                })
        
        # Crear DataFrame del template
        df_template = pd.DataFrame(template_data)
        
        # Exportar a Excel
        with pd.ExcelWriter(archivo_salida, engine='openpyxl', datetime_format='DD/MM/YYYY') as writer:
            df_template.to_excel(writer, index=False, sheet_name='Template')
        
        # Aplicar formato
        wb = load_workbook(archivo_salida)
        ws = wb['Template']
        
        # Estilo del encabezado
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 45
        
        # Formato de fecha para la columna A
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].number_format = 'DD/MM/YYYY'
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        wb.save(archivo_salida)
        
        print(f"✅ Template generado exitosamente: {archivo_salida}")
        print(f"✅ Total de registros: {len(df_template):,}")
        print(f"✅ Días procesados: {len(fechas)}")
        print(f"✅ Tiendas por día: {len(tiendas)}")
        print(f"✅ Cálculo: {len(fechas)} días × {len(tiendas)} tiendas = {len(df_template):,} registros")
        
        return df_template
        
    except Exception as e:
        print(f"❌ Error al generar el template: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

# Ejemplo de uso
if __name__ == "__main__":
    # Configuración
    archivo_entrada = "lista_tiendas.xlsx"  # Tu archivo con las tiendas
    numero_promotor = "30107"  # Número del promotor
    fecha_inicio = "01/10/2025"  # Fecha de inicio
    fecha_fin = "31/10/2025"      # Fecha fin
    
    # Generar el template
    resultado = generar_template_tiendas(
        archivo_entrada=archivo_entrada,
        numero_promotor=numero_promotor,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        archivo_salida='Template.xlsx'
    )
    
    print("\n📊 Muestra de las primeras 10 filas del template:")
    print(resultado.head(10).to_string(index=False))